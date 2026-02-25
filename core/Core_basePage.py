from datetime import datetime
from locators.Objectlocators import Objectlocators
from openpyxl import load_workbook
from pathlib import Path
import sys
import os
from playwright.sync_api import sync_playwright, Page, Browser, BrowserContext
from getgauge.python import data_store
from core.ReportLogger import ReportLogger, log_step, log_complete, log_failed

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

class BasePage:
    """
    Unified Base Class containing all core functionality:
    - Page interactions
    - Login actions
    - Reporting/Logging
    - Assertions
    - Configuration
    """
    # ====================================
    # Playwright Driver Core (Merged)
    # ====================================
    _playwright = None
    _browser: Browser = None
    _context: BrowserContext = None
    _page: Page = None
    
    @classmethod
    def initialize(cls):
        """Initialize Playwright and launch browser"""
        if cls._playwright is None:
            cls._playwright = sync_playwright().start()
            browser_type = cls.get_browser_type()
            headless = cls.is_headless()
            slow_mo = cls.get_slow_mo()
            # Get browser
            launch_args = ["--start-maximized"] if not headless else []
            if browser_type.lower() == 'firefox':
                cls._browser = cls._playwright.firefox.launch(headless=headless,slow_mo=slow_mo)
            elif browser_type.lower() == 'webkit':
                cls._browser = cls._playwright.webkit.launch(headless=headless,slow_mo=slow_mo)
            elif browser_type.lower() == 'edge':
                cls._browser = cls._playwright.chromium.launch(channel="msedge", headless=headless, slow_mo=slow_mo, args=launch_args)
            #elif browser_type.lower() == 'chrome':
            #    cls._browser = cls._playwright.chromium.launch(channel="chrome", headless=headless, slow_mo=slow_mo, args=launch_args)
            else:  # chromium (default)
                cls._browser = cls._playwright.chromium.launch(channel="chrome",headless=headless, slow_mo=slow_mo, args=launch_args)

    @classmethod
    def maximize_window(cls):
        """Maximize browser window - mocked for Playwright compatibility"""
        pass
    
    @classmethod
    def create_context(cls):
        """Create a new browser context"""
        if cls._browser is None:
            cls.initialize()
            cls.maximize_window() 
        # Use no_viewport=True to allow the browser to control the size (needed for maximize)
        cls._context = cls._browser.new_context(no_viewport=True)
        # Set timeouts
        cls._context.set_default_timeout(cls.get_default_timeout())
        cls._context.set_default_navigation_timeout(cls.get_navigation_timeout())
        # Create page
        cls._page = cls._context.new_page()
    
    @classmethod
    def get_page(cls) -> Page:
        """Get the current page instance"""
        return cls._page
    
    @classmethod
    def get_context(cls) -> BrowserContext:
        """Get the current browser context"""
        return cls._context
    
    @classmethod
    def get_browser(cls) -> Browser:
        """Get the browser instance"""
        return cls._browser
    
    @classmethod
    def close_context(cls):
        """Close the current context"""
        if cls._context:
            cls._context.close()
            cls._context = None
            cls._page = None
    
    @classmethod
    def close(cls):
        """Close browser and Playwright"""
        if cls._context:
            cls._context.close()
        if cls._browser:
            cls._browser.close()
        if cls._playwright:
            cls._playwright.stop()

        cls._context = None
        cls._page = None
        cls._browser = None
        cls._playwright = None
    
    @classmethod
    def start_tracing(cls):
        """Start tracing if enabled"""
        if cls.is_tracing_enabled() and cls._context:
            trace_dir = cls.get_trace_dir()
            os.makedirs(trace_dir, exist_ok=True)
            cls._context.tracing.start(screenshots=True, snapshots=True)
    
    @classmethod
    def stop_tracing(cls, name="trace"):
        """Stop tracing and save trace file"""
        if cls.is_tracing_enabled() and cls._context:
            trace_dir = cls.get_trace_dir()
            os.makedirs(trace_dir, exist_ok=True)
            trace_path = os.path.join(trace_dir, f"{name}.zip")
            cls._context.tracing.stop(path=trace_path)
    
    @classmethod
    def take_screenshot(cls, name="screenshot"):
        """Take screenshot and save to file"""
        if cls._page:
            # Use GAUGE_REPORTS_DIR if set, otherwise default to 'reports'
            base_report_dir = os.environ.get("GAUGE_REPORTS_DIR", "reports")
            screenshot_dir = os.path.join(base_report_dir, "screenshots")
            os.makedirs(screenshot_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            screenshot_path = os.path.join(screenshot_dir, f"{name}_{timestamp}.png")
            cls._page.screenshot(path=screenshot_path)
            return screenshot_path
        return None

    # ====================================
    # Configuration Constants
    # ====================================
    # DEFAULT_DATA_FILE = 'td_FrameworkData.xlsx' # Removed hardcoded reference
    DEFAULT_ENV_SHEET = 'Environment'
    DEFAULT_DATA_SHEET = None
    # DEFAULT_CONFIG_PATH = f'data/{DEFAULT_DATA_FILE}' # Removed hardcoded reference

    # ====================================
    # Configuration Cache
    # ====================================
    _config_cache = None
    _config_file = None
    _soft_errors = []
    
    # ====================================
    # Default Locators (Class Level)
    # ====================================
    username_input = Objectlocators.USERNAME_INPUT
    password_input = Objectlocators.PASSWORD_INPUT
    login_button = Objectlocators.LOGIN_BUTTON

    # ====================================
    # Browser Actions
    # ====================================
    
    @classmethod
    def navigate_to(cls, url):
        """Navigate to the specified URL with reporting and error handling"""
        try:
            log_step("Navigate to URL")
            ReportLogger.log_url(url)
            cls.get_page().goto(url)
            log_complete("Navigation successful")
        except Exception as e:
            log_failed(f"Navigation failed: {str(e)}")
            cls.take_screenshot("Error_Navigate")
            assert False, f"Navigation failed: {str(e)}"

    @classmethod
    def get_title(cls):
        """Get the current page title."""
        return cls.get_page().title()

    # ====================================
    # Login Actions (Formerly LoginPage)
    # ====================================

    @classmethod
    def login(cls, username, password):
        """Perform login with reporting and error handling"""
        try:
            log_step("Login Action")
            ReportLogger.log_credentials(username, password) 
            cls.get_page().fill(cls.username_input, username)
            cls.get_page().fill(cls.password_input, password)
            cls.get_page().click(cls.login_button)
            log_complete("Login action performed")
        except Exception as e:
            log_failed(f"Login failed: {str(e)}")
            cls.take_screenshot("Error_Login")
            assert False, f"Login failed: {str(e)}"
            
    # ====================================
    # Assertions (Formerly Assertions class)
    # ====================================
    @classmethod
    def assert_equal(cls, actual, expected, message=""):
        """Assert that actual equals expected."""
        assert actual == expected, f"{message}: Expected '{expected}', got '{actual}'"
    
    @classmethod
    def assert_not_equal(cls, actual, expected, message=""):
        """Assert that actual does not equal expected."""
        assert actual != expected, f"{message}: Values should not be equal - '{actual}'"
    
    @classmethod
    def assert_true(cls, condition, message=""):
        """Assert that condition is True."""
        assert condition, f"{message}: Condition is not True"
    
    @classmethod
    def assert_false(cls, condition, message=""):
        """Assert that condition is False."""
        assert not condition, f"{message}: Condition is not False"
    
    @classmethod
    def assert_contains(cls, text, substring, message=""):
        """Assert that text contains substring."""
        assert substring in text, f"{message}: '{text}' does not contain '{substring}'"
    
    @classmethod
    def assert_not_contains(cls, text, substring, message=""):
        """Assert that text does not contain substring."""
        assert substring not in text, f"{message}: '{text}' should not contain '{substring}'"
    
    @classmethod
    def assert_greater_than(cls, actual, expected, message=""):
        """Assert that actual is greater than expected."""
        assert actual > expected, f"{message}: {actual} is not greater than {expected}"
    
    @classmethod
    def assert_less_than(cls, actual, expected, message=""):
        """Assert that actual is less than expected."""
        assert actual < expected, f"{message}: {actual} is not less than {expected}"
    
    # Soft Assertions
    @classmethod
    def soft_assert_equal(cls, actual, expected, message=""):
        """Soft assert - collects failures instead of failing immediately."""
        if actual != expected:
            cls._soft_errors.append(f"{message}: Expected '{expected}', got '{actual}'")
    
    @classmethod
    def soft_assert_true(cls, condition, message=""):
        """Soft assert for True condition."""
        if not condition:
            cls._soft_errors.append(f"{message}: Condition is not True")
    
    @classmethod
    def soft_assert_contains(cls, text, substring, message=""):
        """Soft assert for substring presence."""
        if substring not in text:
            cls._soft_errors.append(f"{message}: '{text}' does not contain '{substring}'")
    
    @classmethod
    def assert_all_soft(cls):
        """Assert all collected soft assertion failures at once."""
        if cls._soft_errors:
            errors = "\n".join(cls._soft_errors)
            cls._soft_errors = []
            raise AssertionError(f"Soft assertion failures:\n{errors}")
    
    @classmethod
    def clear_soft_errors(cls):
        """Clear all collected soft errors."""
        cls._soft_errors = []
    
    @classmethod
    def get_soft_error_count(cls):
        """Get the count of soft errors."""
        return len(cls._soft_errors)
    
    # Playwright-specific Assertions
    @classmethod
    def assert_element_visible(cls, page, selector, message=""):
        """Assert that an element is visible on the page."""
        assert page.is_visible(selector), f"{message}: Element not visible - {selector}"
    
    @classmethod
    def assert_element_not_visible(cls, page, selector, message=""):
        """Assert that an element is not visible on the page."""
        assert not page.is_visible(selector), f"{message}: Element should not be visible - {selector}"
    
    @classmethod
    def assert_element_enabled(cls, page, selector, message=""):
        """Assert that an element is enabled."""
        assert page.is_enabled(selector), f"{message}: Element not enabled - {selector}"
    
    @classmethod
    def assert_element_disabled(cls, page, selector, message=""):
        """Assert that an element is disabled."""
        assert not page.is_enabled(selector), f"{message}: Element should be disabled - {selector}"
    
    @classmethod
    def assert_text_equals(cls, page, selector, expected_text, message=""):
        """Assert that element's text equals expected text."""
        actual_text = page.text_content(selector)
        assert actual_text == expected_text, f"{message}: Expected '{expected_text}', got '{actual_text}'"
    
    @classmethod
    def assert_text_contains(cls, page, selector, expected_text, message=""):
        """Assert that element's text contains expected text."""
        actual_text = page.text_content(selector)
        assert expected_text in actual_text, f"{message}: '{actual_text}' does not contain '{expected_text}'"
    
    @classmethod
    def assert_url_equals(cls, page, expected_url, message=""):
        """Assert that current URL equals expected URL."""
        actual_url = page.url
        assert actual_url == expected_url, f"{message}: Expected URL '{expected_url}', got '{actual_url}'"
    
    @classmethod
    def assert_url_contains(cls, page, expected_substring, message=""):
        """Assert that current URL contains expected substring."""
        actual_url = page.url
        assert expected_substring in actual_url, f"{message}: URL '{actual_url}' does not contain '{expected_substring}'"
    
    @classmethod
    def assert_title_equals(cls, page, expected_title, message=""):
        """Assert that page title equals expected title."""
        actual_title = page.title()
        assert actual_title == expected_title, f"{message}: Expected title '{expected_title}', got '{actual_title}'"

    # ====================================
    # Configuration (Formerly ConfigManager/EnvConfigReader)
    # ====================================

    @classmethod
    def _load_excel_config(cls):
        """
        Load configuration from Excel file into cache.
        Dynamically discovers the config file if not set via environment variable.
        """
        if cls._config_cache is not None:
            return cls._config_cache

        # Determine config file path
        # Priority 1: Environment Variable
        config_file_env = os.getenv('ENV_CONFIG_FILE')
        base_dir = Path(__file__).parent.parent
        config_path = None

        if config_file_env:
            config_path = base_dir / config_file_env
            if not config_path.exists():
                raise FileNotFoundError(f"Environment config file specified in env var not found: {config_path}")
        else:
            # Priority 2: TEST_DATA_FILE from env property (loaded by Gauge from default.properties)
            # This allows the test data file (which contains Environment sheet) to be single source of truth
            data_dir = base_dir / 'data'
            test_data_file = os.getenv('TEST_DATA_FILE')
            
            if data_dir.exists():
                # If TEST_DATA_FILE is defined in properties, try to use it as config file
                if test_data_file:
                    possible_path = data_dir / test_data_file
                    if possible_path.exists():
                        # Check if it has Environment sheet, or just assume it does if user configured it
                         try:
                            wb = load_workbook(possible_path, read_only=True, data_only=True)
                            if 'Environment' in wb.sheetnames:
                                config_path = possible_path
                                wb.close()
                            else:
                                wb.close()
                         except Exception:
                             pass

                # If config path still not found, try discovery
                if not config_path:
                    # Find any xlsx with 'Environment' sheet
                    xlsx_files = list(data_dir.glob('*.xlsx'))
                    for file_path in xlsx_files:
                        if file_path.name.startswith('~$'): continue # Skip temporary files
                        try:
                            # Use read_only=True for performance when just checking existence
                            wb = load_workbook(file_path, read_only=True, data_only=True)
                            if 'Environment' in wb.sheetnames:
                                config_path = file_path
                                wb.close()
                                break
                            wb.close()
                        except Exception:
                            continue
        
        if not config_path or not config_path.exists():
            raise FileNotFoundError(f"Could not find a valid configuration Excel file in {base_dir}/data/ with an 'Environment' sheet.")

        config_sheet = os.getenv('ENV_CONFIG_SHEET', cls.DEFAULT_ENV_SHEET)
        
        # Load Excel file
        workbook = load_workbook(config_path, data_only=True)
        if config_sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet '{config_sheet}' not found in {config_path}")
        worksheet = workbook[config_sheet]
        
        # Parse configuration into dictionary
        config = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):  # Skip header row
            if row[0]:  # If property name exists
                property_name = str(row[0]).strip()
                property_value = row[1] if row[1] is not None else ""
                config[property_name] = str(property_value).strip() if property_value else ""
        workbook.close()
        cls._config_cache = config
        cls._config_file = str(config_path)
        return config
    
    @classmethod
    def _get_from_excel(cls, key, default=None):
        """Get configuration value from Excel."""
        config = cls._load_excel_config()
        return config.get(key, default)
    
    @classmethod
    def _get_config(cls, key, default=None):
        """Get configuration value (env var takes priority over Excel)."""
        return os.getenv(key) or cls._get_from_excel(key, default)
    
    @classmethod
    def _get_config_int(cls, key, default=0):
        """Get configuration value as integer."""
        env_value = os.getenv(key)
        if env_value:
            try:
                return int(env_value)
            except (ValueError, TypeError):
                pass
        excel_value = cls._get_from_excel(key, default)
        try:
            return int(excel_value)
        except (ValueError, TypeError):
            return default
    
    @classmethod
    def _get_config_bool(cls, key, default=False):
        """Get configuration value as boolean."""
        env_value = os.getenv(key)
        if env_value is not None:
            return env_value.lower() in ('true', 'yes', '1', 'on')
        excel_value = cls._get_from_excel(key, default)
        if isinstance(excel_value, bool):
            return excel_value
        if excel_value is None:
            return default
        return str(excel_value).lower() in ('true', 'yes', '1', 'on')
    
    @classmethod
    def reload_config(cls):
        """Force reload configuration from Excel file."""
        cls._config_cache = None
        return cls._load_excel_config()
    
    @classmethod
    def get_all_config(cls):
        """Get all configuration as dictionary."""
        return cls._load_excel_config().copy()
    
    # --- Public Config Getters ---
    @classmethod
    def get_config(cls, key, default=None):
        return cls._get_config(key, default)

    @classmethod
    def get_config_int(cls, key, default=0):
        return cls._get_config_int(key, default)

    @classmethod
    def get_config_bool(cls, key, default=False):
        return cls._get_config_bool(key, default)

    @classmethod
    def get_app_url(cls):
        url = cls._get_config('APP_URL')
        if not url:
            raise ValueError("APP_URL not found in Environment configuration")
        return url
    
    @classmethod
    def get_app_name(cls):
        name = cls._get_config('APP_NAME')
        if not name:
            return 'Test Application' # Fallback only if strictly needed, or raise Error. User asked to parameterize. I will keep fallback but ensure it reads config. 
            # Actually user said "Can you parameter values... from excel". 
            # I will maintain the existing pattern of _get_config which prefers Env/Excel. 
            # But I will remove the explicit string default in the argument to emphasize explicit config.
        return name 

    @classmethod
    def get_environment(cls):
        env = cls._get_config('ENVIRONMENT')
        if not env:
            raise ValueError("ENVIRONMENT not configured in Environment properties")
        return env
    
    @classmethod
    def get_browser_type(cls):
        browser = cls._get_config('BROWSER')
        if not browser:
             raise ValueError("BROWSER not configured in Environment properties")
        return browser
    
    @classmethod
    def is_headless(cls):
        val = cls._get_config('HEADLESS') # Get raw to check existence
        if val is None or val == "":
             # Default to True/False if really missing, but user wants parameterization.
             # Let's trust _get_config_bool but ensure we aren't hiding a missing value with a default arg
             # Actually safer to let _get_config_bool handle parsing, but we pass no default?
             # _get_config_bool uses default. 
             # Let's assume strictness.
             raise ValueError("HEADLESS not configured in Environment properties")
        return cls._get_config_bool('HEADLESS', False)
    
    @classmethod
    def get_slow_mo(cls):
        val = cls._get_config('SLOW_MO')
        if val is None:
             # If missing in strict mode, raise error or return 0? 
             # User asked to parameter values.
             # I'll rely on int conversion which handles 0 fine, but None means missing.
             # I previously updated Excel with 1000 default if missing.
             val = cls._get_config_int('SLOW_MO', None)
             if val is None: raise ValueError("SLOW_MO not configured")
             return val
        return cls._get_config_int('SLOW_MO', 0)
    
    @classmethod
    def get_viewport_width(cls):
        val = cls._get_config_int('VIEWPORT_WIDTH', None)
        if val is None: raise ValueError("VIEWPORT_WIDTH not configured")
        return val
    
    @classmethod
    def get_viewport_height(cls):
        val = cls._get_config_int('VIEWPORT_HEIGHT', None)
        if val is None: raise ValueError("VIEWPORT_HEIGHT not configured")
        return val
    
    @classmethod
    def get_default_timeout(cls):
        val = cls._get_config_int('DEFAULT_TIMEOUT', None)
        if val is None: raise ValueError("DEFAULT_TIMEOUT not configured")
        return val
    
    @classmethod
    def get_navigation_timeout(cls):
        val = cls._get_config_int('NAVIGATION_TIMEOUT', None)
        if val is None: raise ValueError("NAVIGATION_TIMEOUT not configured")
        return val
    
    @classmethod
    def get_action_timeout(cls):
        val = cls._get_config_int('ACTION_TIMEOUT', None)
        if val is None: raise ValueError("ACTION_TIMEOUT not configured")
        return val
    
    @classmethod
    def get_visibility_timeout(cls):
        val = cls._get_config_int('VISIBILITY_TIMEOUT', None)
        if val is None: raise ValueError("VISIBILITY_TIMEOUT not configured")
        return val
    
    @classmethod
    def screenshot_on_failure(cls):
        val = cls._get_config_bool('SCREENSHOT_ON_FAILURE', None)
        if val is None: raise ValueError("SCREENSHOT_ON_FAILURE not configured")
        return val
    
    @classmethod
    def is_tracing_enabled(cls):
        val = cls._get_config_bool('ENABLE_TRACING', None)
        if val is None: raise ValueError("ENABLE_TRACING not configured")
        return val
    
    @classmethod
    def get_trace_dir(cls):
        val = cls._get_config('TRACE_DIR')
        if not val: raise ValueError("TRACE_DIR not configured")
        return val
    
    @classmethod
    def get_test_data_file(cls):
        val = cls._get_config('TEST_DATA_FILE')
        if not val: raise ValueError("TEST_DATA_FILE not configured")
        return val

    @classmethod
    def get_env_config_sheet(cls):
        val = cls._get_config('ENV_CONFIG_SHEET')
        if not val: raise ValueError("ENV_CONFIG_SHEET not configured")
        return val
    
    @classmethod
    def get_test_data_sheet(cls):
        sheet = cls._get_config('TEST_DATA_SHEET')
        if not sheet:
             if cls.DEFAULT_DATA_SHEET:
                 return cls.DEFAULT_DATA_SHEET
             raise ValueError("TEST_DATA_SHEET not configured in Environment properties (Excel) and no default is set.")
        return sheet
    
    @classmethod
    def get_retry_count(cls):
        val = cls._get_config_int('RETRY_COUNT', None)
        if val is None: return 0 # Retry count can be optional/0
        return val
    
    @classmethod
    def get_log_level(cls):
        return cls._get_config('LOG_LEVEL') or 'INFO' # Info is safe default
    
    @classmethod
    def get_base_path(cls):
        val = cls._get_config('BASE_PATH')
        if not val: raise ValueError("BASE_PATH not configured")
        return val

    @classmethod
    def get_test_id_columns(cls):
        """Get list of Test ID column names from config."""
        columns = cls._get_config('TEST_ID_COLUMNS')
        if not columns: raise ValueError("TEST_ID_COLUMNS not configured")
        return [c.strip() for c in columns.split(',') if c.strip()]

    @staticmethod
    def get_test_id_from_tags():
        """
        Extract Test ID directly from scenario tags.
        Assumes the tag IS the Test ID.
        """
        scenario_tags = data_store.scenario.get('tags', [])
        # Return the first tag as the ID, or filter if needed. 
        # Usually user might have multiple tags. 
        # If we assume 'Tags' column in excel matches one of these tags.
        if scenario_tags:
            return scenario_tags[0] # Return the first tag as the ID
        return None

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def get_sheet_for_test_id(test_id):
    """Dynamically find which sheet contains the given Test ID (Tags)"""
    if not test_id:
        return None
    test_data_file = BasePage.get_test_data_file()
    try:
        from core.TestDataManager import ExcelReader  # Local import to avoid circular dependency
        excel = ExcelReader(test_data_file)
        all_sheets = excel.get_all_sheet_names()
        for sheet_name in all_sheets:
            if sheet_name.lower() in ['environment', 'config', 'settings']:
                continue
            try:
                # Look for columns dynamically
                target_columns = BasePage.get_test_id_columns()
                test_data = excel.get_test_data_by_id(sheet_name, test_id, target_columns)
                if test_data:
                    excel.close()
                    return sheet_name
            except Exception:
                continue
        excel.close()
        return BasePage.get_test_data_sheet()
    except Exception as e:
        print(f"Error searching for Test ID {test_id}: {str(e)}")
        return BasePage.get_test_data_sheet()

def get_all_test_ids_from_excel():
    """Dynamically discover all Test IDs across all sheets"""
    test_data_file = BasePage.get_test_data_file()
    test_id_mapping = {}
    try:
        from core.TestDataManager import ExcelReader  # Local import to avoid circular dependency
        excel = ExcelReader(test_data_file)
        all_sheets = excel.get_all_sheet_names()
        for sheet_name in all_sheets:
            if sheet_name.lower() in ['environment', 'config', 'settings']:
                continue
            try:
                all_data = excel.get_all_data(sheet_name)
                # Strict columns
                target_columns = BasePage.get_test_id_columns()
                for row in all_data:
                    for key in target_columns:
                        if key in row and row[key]:
                            test_id_mapping[row[key]] = sheet_name
                            break
            except Exception:
                continue
        excel.close()
        return test_id_mapping
    except Exception as e:
        print(f"Error discovering Test IDs: {str(e)}")
        return {}
