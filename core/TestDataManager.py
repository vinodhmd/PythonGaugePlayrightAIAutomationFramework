"""
Test Data Manager - Unified test data management
Handles both Excel reading and automatic data loading based on scenario tags
"""
from openpyxl import load_workbook
from core.Core_basePage import BasePage, get_sheet_for_test_id
from getgauge.python import data_store
import os

class TestDataManager:
    """
    Unified class for Excel reading and automatic test data management.
    Combines functionality of ExcelReader and TestDataManager.
    """ 
    # ========================================================================
    # EXCEL READING METHODS (Low-level operations)
    # ========================================================================
    def __init__(self, file_name): 
        """ Initialize Excel reader with file name.
        Args:file_name: Name of the Excel file (will be looked up in data/ folder)
        """
        data_folder = os.path.join(os.path.dirname(os.path.dirname(__file__)),'data')
        self.file_path = os.path.join(data_folder, file_name)
        self.workbook = load_workbook(self.file_path)
    
    def get_row_data(self, sheet_name, row_number):
        """Get data from a specific row in the Excel sheet.
        Args: sheet_name: Name of the sheet
            row_number: Row number to read (1-indexed)       
        Returns: dict: Dictionary with column headers as keys and row values as values
        """
        sheet = self.workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        row_data = [cell.value for cell in sheet[row_number]]
        return dict(zip(headers, row_data))
    
    def get_test_data_by_id(self, sheet_name, test_id, test_id_columns):
        """
        Get test data by searching for test ID in specified columns.
        Args:
            sheet_name: Name of the sheet to search in
            test_id: Test ID to search for (e.g., 'TC001', 'PAY001')
            test_id_columns: List of column names to search in (e.g., ['Tags'])
        Returns:
            dict: Dictionary with test data, or None if not found
        """
        sheet = self.workbook[sheet_name]
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        # Find which columns to search in
        search_column_indices = []
        for col_name in test_id_columns:
            if col_name in headers:
                search_column_indices.append(headers.index(col_name))
        if not search_column_indices:
            return None
        # Search for test_id in the specified columns
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for col_idx in search_column_indices:
                if col_idx < len(row) and row[col_idx] == test_id:
                    # Found the test ID, return the row data as dictionary
                    return dict(zip(headers, row))
        return None
    
    def get_all_sheet_names(self):
        """
        Get all sheet names in the workbook.
        Returns:
            list: List of sheet names
        """
        return self.workbook.sheetnames
    
    def close(self):
        """Close the workbook."""
        self.workbook.close()
    
    # ========================================================================
    # AUTOMATIC TEST DATA MANAGEMENT (High-level Gauge integration)
    # ========================================================================
    
    @classmethod
    def load_test_data(cls):
        """
        Automatically load test data based on scenario tags.
        This should be called in hooks before_scenario.
        Returns:
            dict: Test data for the current scenario, or None if no matching tag found
        """
        # Get test ID from scenario tags
        test_id = BasePage.get_test_id_from_tags()
        if not test_id:
            # No test ID found in tags, return None
            return None
        # Get test data file and sheet
        test_data_file = BasePage.get_test_data_file()
        test_data_sheet = get_sheet_for_test_id(test_id)
        if not test_data_sheet:
            return None
        # Fetch test data from Excel using instance methods
        excel = cls(test_data_file)
        test_id_columns = BasePage.get_test_id_columns()
        test_data = excel.get_test_data_by_id(test_data_sheet, test_id, test_id_columns)
        excel.close()
        # Store in data_store for access across steps
        if test_data:
            data_store.scenario['test_data'] = test_data
            data_store.scenario['test_id'] = test_id
            data_store.scenario['test_sheet'] = test_data_sheet
        return test_data
    
    @classmethod
    def get_test_data(cls):
        """
        Get the test data for the current scenario.
        Returns:
            dict: Test data dictionary, or None if not loaded
        """
        return data_store.scenario.get('test_data', None)
    
    @classmethod
    def get_test_id(cls):
        """
        Get the test ID for the current scenario.
        Returns:
            str: Test ID (e.g., 'PAY001', 'TC001'), or None
        """
        return data_store.scenario.get('test_id', None)
    
    @classmethod
    def get_test_sheet(cls):
        """
        Get the sheet name where test data was loaded from.
        Returns:
            str: Sheet name (e.g., 'EmployeeCreation', 'LoginData'), or None
        """
        return data_store.scenario.get('test_sheet', None)
    
    @classmethod
    def get_value(cls, key, default=None):
        """
        Get a specific value from test data.
        Args:
            key: The key to retrieve (e.g., 'Username', 'Password', 'Url')
            default: Default value if key not found
        Returns:
            The value for the key, or default if not found
        """
        test_data = cls.get_test_data()
        if test_data:
            return test_data.get(key, default)
        return default
    
    @classmethod
    def has_test_data(cls):
        """
        Check if test data is loaded for the current scenario.
        Returns:
            bool: True if test data is available, False otherwise
        """
        return cls.get_test_data() is not None
    
    @classmethod
    def clear(cls):
        """
        Clear test data for the current scenario.
        This should be called in hooks after_scenario.
        """
        if 'test_data' in data_store.scenario:
            del data_store.scenario['test_data']
        if 'test_id' in data_store.scenario:
            del data_store.scenario['test_id']
        if 'test_sheet' in data_store.scenario:
            del data_store.scenario['test_sheet']

    @classmethod
    def get_test_data_file(cls):
        """Delegate to BasePage to get configured test data file"""
        return BasePage.get_test_data_file()

# Backward compatibility alias
ExcelReader = TestDataManager
