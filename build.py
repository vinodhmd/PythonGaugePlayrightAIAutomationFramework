import os
import shutil
import subprocess
import sys
import argparse
import openpyxl
from pathlib import Path

def clean():
    """Removes build artifacts and temporary files."""
    print("Cleaning build artifacts...")
    directories_to_remove = ["reports", ".gauge"]
    
    for directory in directories_to_remove:
        if os.path.exists(directory):
            try:
                shutil.rmtree(directory)
                print(f"Removed directory: {directory}")
            except Exception as e:
                print(f"Error removing {directory}: {e}")
                
    # Clean pycache
    for root, dirs, files in os.walk("."):
        for dir in dirs:
            if dir == "__pycache__":
                try:
                    shutil.rmtree(os.path.join(root, dir))
                    print(f"Removed __pycache__ in {root}")
                except Exception as e:
                    print(f"Error removing __pycache__ in {root}: {e}")
        for file in files:
            if file.endswith(".pyc"):
                 try:
                    os.remove(os.path.join(root, file))
                    print(f"Removed .pyc file: {os.path.join(root, file)}")
                 except Exception as e:
                    print(f"Error removing {file}: {e}")

def install():
    """Installs dependencies from requirements.txt."""
    print("Installing dependencies...")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-r", "requirements.txt"])
        print("Dependencies installed successfully.")
    except subprocess.CalledProcessError as e:
        print(f"Error installing dependencies: {e}")
        sys.exit(1)

def get_excel_config():
    """Reads configuration from Excel file."""
    # Try to read TEST_DATA_FILE from env/default/default.properties
    config_excel_name = "td_FrameworkData.xlsx" # Default fallback
    default_props_path = Path("env/default/default.properties")
    
    if default_props_path.exists():
        try:
            with open(default_props_path, 'r') as f:
                for line in f:
                    if line.strip().startswith("TEST_DATA_FILE"):
                        parts = line.split("=")
                        if len(parts) > 1:
                            val = parts[1].strip()
                            if val:
                                config_excel_name = val
                                break
        except Exception:
            pass

    config_path = Path(f"data/{config_excel_name}")
    sheet_name = "Environment"
    config = {}
    
    if not config_path.exists():
        print(f"Warning: Config file not found at {config_path}. Using defaults.")
        return config
    try:
        workbook = openpyxl.load_workbook(config_path, data_only=True)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    key = str(row[0]).strip()
                    value = row[1] if row[1] is not None else ""
                    config[key] = str(value).strip()
        workbook.close()
    except Exception as e:
        print(f"Error reading Excel config: {e}")
    return config

def setup_environment(env_name):
    """Creates environment folder if it doesn't exist by copying default."""
    if not env_name or env_name.lower() == "default":
        return
    src = Path("env/default")
    dst = Path(f"env/{env_name}")
    if not src.exists():
        print("Error: env/default source directory not found.")
        return
    if dst.exists():
        shutil.rmtree(dst)
    try:
        shutil.copytree(src, dst)
        print(f"Created environment directory: {dst}")
    except Exception as e:
        print(f"Error creating environment directory: {e}")

def run():
    """Runs the Gauge specs using environment from Excel."""
    print("Reading configuration...")
    config = get_excel_config()
    env_name = config.get("ENVIRONMENT", "default")
    app_name = config.get("APP_NAME", "Gauge Project")
    print(f"Configuration loaded: Environment='{env_name}', AppName='{app_name}'")    
    setup_environment(env_name)
    print(f"Running Gauge specs in environment: {env_name}...")
    env_vars = os.environ.copy()
    env_vars["GAUGE_PROJECT_NAME"] = app_name
    try:
        cmd = f"gauge run specs --env {env_name}"
        subprocess.check_call(cmd, shell=True, env=env_vars)
        print("Tests execution completed.")
    except subprocess.CalledProcessError as e:
        print(f"Tests failed with exit code {e.returncode}.")
        sys.exit(e.returncode)

def main():
    parser = argparse.ArgumentParser(description='Build and test the Gauge framework')
    parser.add_argument('--skip-install', action='store_true', 
                        help='Skip dependency installation')
    parser.add_argument('--skip-clean', action='store_true', 
                        help='Skip cleaning build artifacts')
    args = parser.parse_args()
    if not args.skip_clean:
        clean()
    else:
        print("Skipping clean step...")
    if not args.skip_install:
        install()
    else:
        print("Skipping dependency installation...")
    
    #run()

if __name__ == "__main__":
    main()
