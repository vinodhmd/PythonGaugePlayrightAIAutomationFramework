import os
import openpyxl
from pathlib import Path
import sys

# Add project root to path before importing core modules
yml_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(yml_dir)
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from core.TestDataManager import TestDataManager
def load_env_context():
    """
    Reads configuration from the configured Excel file (defined in default.properties)
    and returns a dictionary of environment variables.
    """
    yml_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(yml_dir)
    
    # 1. Read default.properties to find TEST_DATA_FILE
    props_path = os.path.join(project_root, 'env', 'default', 'default.properties')
    # test_data_file = 'td_FrameworkData.xlsx' # Default
    test_data_file = TestDataManager.get_test_data_file()
    
    if os.path.exists(props_path):
        try:
            with open(props_path, 'r') as f:
                for line in f:
                    if line.strip().startswith('TEST_DATA_FILE'):
                        parts = line.split('=')
                        if len(parts) > 1:
                            val = parts[1].strip()
                            if val:
                                test_data_file = val
                                break
        except Exception as e:
            print(f"Warning: Failed to read default.properties: {e}")

    # 2. Read the Excel File
    excel_path = os.path.join(project_root, 'data', test_data_file)
    env_vars = {}
    
    if os.path.exists(excel_path):
        try:
            # Use data_only=True to get values instead of formulas
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            if 'Environment' in wb.sheetnames:
                ws = wb['Environment']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        key = str(row[0]).strip()
                        value = str(row[1]).strip() if row[1] is not None else ""
                        env_vars[key] = value
            wb.close()
            print(f"Loaded {len(env_vars)} environment variables from {test_data_file}")
        except Exception as e:
            print(f"Warning: Failed to read Excel config {test_data_file}: {e}")
    else:
        print(f"Warning: Test data file not found: {excel_path}")
        
    return env_vars
